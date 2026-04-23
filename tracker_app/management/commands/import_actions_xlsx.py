from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from openpyxl import load_workbook

from tracker_app.models import Action, ActionStatus, Objective


def normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def to_text(value):
    if value is None:
        return ""
    return str(value).strip()


class Command(BaseCommand):
    help = "Import Action rows from an Excel .xlsx file"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_file", type=str, help="Path to the .xlsx file")
        parser.add_argument(
            "--sheet",
            type=str,
            default=None,
            help="Worksheet name (defaults to active sheet)",
        )
        parser.add_argument(
            "--update",
            action="store_true",
            help="Update existing Action rows with the same title",
        )

    def handle(self, *args, **options):
        xlsx_file = Path(options["xlsx_file"])
        if not xlsx_file.exists():
            raise CommandError(f"File not found: {xlsx_file}")
        if xlsx_file.suffix.lower() != ".xlsx":
            raise CommandError("Only .xlsx files are supported")

        workbook = load_workbook(filename=xlsx_file, read_only=True, data_only=True)
        if options["sheet"]:
            sheet_name = options["sheet"]
            if sheet_name not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames)
                raise CommandError(
                    f"Sheet '{sheet_name}' not found. Available sheets: {available}"
                )
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active

        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            raise CommandError("Excel file is empty")

        headers = [normalize_header(h) for h in raw_headers]
        required_headers = {"title"}
        missing = sorted(required_headers - set(headers))
        if missing:
            raise CommandError(
                "Missing required columns: " + ", ".join(missing)
            )

        objective_column = None
        if "objective" in headers:
            objective_column = "objective"
        elif "objective_title" in headers:
            objective_column = "objective_title"
        else:
            raise CommandError("Missing required column: objective or objective_title")

        status_aliases = {
            "inactive": ActionStatus.NOT_STARTED,
            "not_started": ActionStatus.NOT_STARTED,
            "not started": ActionStatus.NOT_STARTED,
            "in_progress": ActionStatus.IN_PROGRESS,
            "in progress": ActionStatus.IN_PROGRESS,
            "completed": ActionStatus.COMPLETED,
        }
        valid_status_values = {choice.value for choice in ActionStatus}

        created = 0
        updated = 0
        skipped = 0
        errors = 0

        for row_index, row in enumerate(rows, start=2):
            data = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    data[headers[i]] = cell

            title = to_text(data.get("title"))
            objective_title = to_text(data.get(objective_column))

            if not title:
                continue
            if not objective_title:
                self.stderr.write(
                    f"Row {row_index}: objective is empty for title '{title}', skipped"
                )
                errors += 1
                continue

            try:
                objective = Objective.objects.get(title=objective_title)
            except Objective.DoesNotExist:
                self.stderr.write(
                    f"Row {row_index}: objective '{objective_title}' not found, skipped"
                )
                errors += 1
                continue

            raw_status = to_text(data.get("status"))
            if raw_status:
                status_key = raw_status.lower().strip()
                status = status_aliases.get(status_key, raw_status)
                if status not in valid_status_values:
                    self.stderr.write(
                        f"Row {row_index}: invalid status '{raw_status}', defaulted to NOT_STARTED"
                    )
                    status = ActionStatus.NOT_STARTED
            else:
                status = ActionStatus.NOT_STARTED

            defaults = {
                "objective": objective,
                "small_description": to_text(data.get("small_description")),
                "small_description_ga": to_text(data.get("small_description_ga")) or None,
                "description": to_text(data.get("description")),
                "description_ga": to_text(data.get("description_ga")),
                "update": to_text(data.get("update")),
                "update_ga": to_text(data.get("update_ga")),
                "status": status,
            }

            action, was_created = Action.objects.get_or_create(title=title, defaults=defaults)
            if was_created:
                created += 1
                continue

            if options["update"]:
                for field_name, value in defaults.items():
                    setattr(action, field_name, value)
                action.save()
                updated += 1
            else:
                skipped += 1

        self.stdout.write(
            self.style.SUCCESS(
                "Import finished from sheet "
                f"'{worksheet.title}'. Created: {created}, Updated: {updated}, "
                f"Skipped: {skipped}, Errors: {errors}"
            )
        )
